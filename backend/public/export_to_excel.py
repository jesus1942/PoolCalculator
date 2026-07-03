#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
import sys
import json
from datetime import datetime
import os
from pathlib import Path
import math
import tempfile
from PIL import Image, ImageDraw, ImageFont

def normalize(value):
    return str(value or '').strip().lower()

def sum_plumbing_items(items, keywords=()):
    total = 0
    for item in items:
        name = normalize(item.get('name', ''))
        if keywords and not any(keyword in name for keyword in keywords):
            continue
        total += float(item.get('quantity', 0) or 0)
    return total

def dominant_pipe_diameter(items):
    ranked = {}
    for item in items:
        name = normalize(item.get('name', ''))
        if 'caño' not in name and 'tubo' not in name:
            continue
        diameter = (item.get('diameter') or '40mm').upper()
        ranked[diameter] = ranked.get(diameter, 0) + float(item.get('quantity', 0) or 0)
    if not ranked:
        return '40MM'
    return max(ranked.items(), key=lambda pair: pair[1])[0]

def safe_float(value, default=0):
    try:
        if value in (None, ''):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default

def safe_int(value, default=0):
    try:
        if value in (None, ''):
            return default
        return int(round(float(value)))
    except (TypeError, ValueError):
        return default

def bags_from_kg(weight_kg, bag_weight_kg=25):
    kg = safe_float(weight_kg, 0)
    if kg <= 0:
        return 0
    return math.ceil(kg / bag_weight_kg)

def calculate_tiles_for_dimension(dimension_length_m):
    tile_size = 0.5
    joint_size = 0.003
    effective_tile_size = tile_size + joint_size

    full_tiles_count = math.floor(dimension_length_m / effective_tile_size)
    remaining_space = dimension_length_m - (full_tiles_count * effective_tile_size)

    if remaining_space <= joint_size:
        return full_tiles_count
    return full_tiles_count + 1

def get_side_config(tile_calculation, side):
    return (tile_calculation.get(side, {}) or {}) if tile_calculation else {}

def has_side_tiles(tile_calculation, side):
    side_config = get_side_config(tile_calculation, side)
    return bool(side_config.get('firstRingType')) or safe_int(side_config.get('rows')) > 0 or bool(side_config.get('selectedTileId'))

def get_excel_layout_data(pool, tile_calculation):
    tile_calculation = tile_calculation or {}

    west = get_side_config(tile_calculation, 'west')
    east = get_side_config(tile_calculation, 'east')
    north = get_side_config(tile_calculation, 'north')
    south = get_side_config(tile_calculation, 'south')

    # Mapeo fijo según TileEditor:
    # west = Skimmer (Lateral Izquierdo)
    # east = Escalera (Lateral Derecho)
    # north = Izquierdo (Cabecera Superior)
    # south = Derecho (Cabecera Inferior)
    stairs_count = safe_int(pool.get('stairsCount'))
    if stairs_count <= 0 and pool.get('hasStairsOnly'):
        stairs_count = 4
    if stairs_count <= 0 and has_side_tiles(tile_calculation, 'east'):
        stairs_count = safe_int(east.get('rows'), 0)

    skimmer_count = safe_int(pool.get('skimmersCount'))
    if skimmer_count <= 0 and has_side_tiles(tile_calculation, 'west'):
        skimmer_count = max(1, safe_int(west.get('rows'), 1))

    return {
        'stairs': stairs_count,
        'skimmer': skimmer_count,
        'left_head': 1 if has_side_tiles(tile_calculation, 'north') else 0,
        'right_head': 1 if has_side_tiles(tile_calculation, 'south') else 0,
        'west_rows': safe_int(west.get('rows')),
        'east_rows': safe_int(east.get('rows')),
        'north_rows': safe_int(north.get('rows')),
        'south_rows': safe_int(south.get('rows')),
    }

def get_side_layers(tile_calculation):
    tile_calculation = tile_calculation or {}
    return {
        'left': max(0, safe_int(get_side_config(tile_calculation, 'north').get('rows'))),
        'right': max(0, safe_int(get_side_config(tile_calculation, 'south').get('rows'))),
        'top': max(0, safe_int(get_side_config(tile_calculation, 'east').get('rows'))),
        'bottom': max(0, safe_int(get_side_config(tile_calculation, 'west').get('rows'))),
    }

def fit_text(draw, text, font, x, y, fill):
    draw.text((x, y), text, font=font, fill=fill)

def load_font(size, bold=False):
    candidates = [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf' if bold else '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf' if bold else '/usr/share/fonts/dejavu/DejaVuSans.ttf',
    ]
    for path in candidates:
        if os.path.exists(path):
            return ImageFont.truetype(path, size=size)
    return ImageFont.load_default()

def create_tile_plan_image(project_data):
    pool = project_data.get('pool', {}) or {}
    tile_calculation = project_data.get('tileCalculation', {}) or {}
    layers = get_side_layers(tile_calculation)

    img_w, img_h = 2800, 1800
    img = Image.new('RGB', (img_w, img_h), '#f8fafc')
    draw = ImageDraw.Draw(img)
    title_font = load_font(48, bold=True)
    section_font = load_font(30, bold=True)
    body_font = load_font(24)

    panel_margin = 70
    info_panel_w = 520
    drawing_x0 = panel_margin
    drawing_y0 = panel_margin
    drawing_w = img_w - (panel_margin * 2) - info_panel_w - 40
    drawing_h = img_h - (panel_margin * 2)

    # Igual que PoolVisualizationCanvas:
    # largo = horizontal, ancho = vertical
    # west/skimmer = izquierda, east/escalera = derecha
    # north = arriba, south = abajo
    pool_length = max(1.0, safe_float(pool.get('length'), 8))
    pool_width = max(1.0, safe_float(pool.get('width'), 4))
    ext_skimmer = layers['bottom']
    ext_escalera = layers['top']
    ext_superior = layers['left']
    ext_inferior = layers['right']

    total_width = pool_length + 0.5 * (ext_skimmer + ext_escalera)
    total_height = pool_width + 0.5 * (ext_superior + ext_inferior)
    scale = min((drawing_w - 180) / total_width, (drawing_h - 180) / total_height)

    outer_w = total_width * scale
    outer_h = total_height * scale
    origin_x = drawing_x0 + (drawing_w - outer_w) / 2
    origin_y = drawing_y0 + (drawing_h - outer_h) / 2

    pool_x = origin_x + ext_skimmer * 0.5 * scale
    pool_y = origin_y + ext_superior * 0.5 * scale
    pool_w = pool_length * scale
    pool_h = pool_width * scale
    tile_px = 0.5 * scale

    # Marco y título
    draw.rounded_rectangle((20, 20, img_w - 20, img_h - 20), radius=24, outline='#cbd5e1', width=3, fill='#ffffff')
    fit_text(draw, f"Plano de Losetas | {project_data.get('projectName', pool.get('name', 'Piscina'))}", title_font, 60, 36, '#0f172a')

    # Área exterior
    draw.rounded_rectangle((drawing_x0, drawing_y0, drawing_x0 + drawing_w, drawing_y0 + drawing_h), radius=24, fill='#e2e8f0', outline='#94a3b8', width=3)

    def tile_colors(side_config):
        first_ring = normalize(side_config.get('firstRingType'))
        if first_ring == 'lomo_ballena':
            return '#fbbf24', '#fef3c7'
        if first_ring:
            return '#cbd5e1', '#f8fafc'
        return '#dbeafe', '#eff6ff'

    side_configs = {
        'north': get_side_config(tile_calculation, 'north'),
        'south': get_side_config(tile_calculation, 'south'),
        'east': get_side_config(tile_calculation, 'east'),
        'west': get_side_config(tile_calculation, 'west'),
    }

    # Vereda continua
    draw.rounded_rectangle(
        (pool_x - ext_skimmer * tile_px, pool_y - ext_superior * tile_px, pool_x + pool_w + ext_escalera * tile_px, pool_y + pool_h + ext_inferior * tile_px),
        radius=22,
        fill='#e2e8f0',
        outline='#94a3b8',
        width=3,
    )

    h_cells = max(1, int(round(pool_length / 0.5)))
    v_cells = max(1, int(round(pool_width / 0.5)))
    cell_w = pool_w / h_cells
    cell_h = pool_h / v_cells

    def draw_top_or_bottom_band(y, rows, side_key):
        if rows <= 0:
            return
        primary, secondary = tile_colors(side_configs[side_key])
        for row in range(rows):
            fill = primary if row == 0 and side_configs[side_key].get('firstRingType') else secondary
            ry = y + row * tile_px
            for i in range(h_cells):
                rx = pool_x + i * cell_w
                draw.rectangle((rx, ry, rx + cell_w, ry + tile_px), fill=fill, outline='#94a3b8', width=2)

    def draw_left_or_right_band(x, cols, side_key):
        if cols <= 0:
            return
        primary, secondary = tile_colors(side_configs[side_key])
        for col in range(cols):
            fill = primary if col == 0 and side_configs[side_key].get('firstRingType') else secondary
            rx = x + col * tile_px
            for i in range(v_cells):
                ry = pool_y + i * cell_h
                draw.rectangle((rx, ry, rx + tile_px, ry + cell_h), fill=fill, outline='#94a3b8', width=2)

    def draw_corner_block(x, y, cols, rows):
        if cols <= 0 or rows <= 0:
            return
        corner_fill = '#fde68a' if any(normalize(side_configs[k].get('firstRingType')) == 'lomo_ballena' for k in side_configs) else '#f8fafc'
        for cx in range(cols):
            for cy in range(rows):
                rx = x + cx * tile_px
                ry = y + cy * tile_px
                draw.rectangle((rx, ry, rx + tile_px, ry + tile_px), fill=corner_fill, outline='#94a3b8', width=2)

    draw_top_or_bottom_band(pool_y - ext_superior * tile_px, ext_superior, 'north')
    draw_top_or_bottom_band(pool_y + pool_h, ext_inferior, 'south')
    draw_left_or_right_band(pool_x - ext_skimmer * tile_px, ext_skimmer, 'west')
    draw_left_or_right_band(pool_x + pool_w, ext_escalera, 'east')

    draw_corner_block(pool_x - ext_skimmer * tile_px, pool_y - ext_superior * tile_px, ext_skimmer, ext_superior)
    draw_corner_block(pool_x + pool_w, pool_y - ext_superior * tile_px, ext_escalera, ext_superior)
    draw_corner_block(pool_x - ext_skimmer * tile_px, pool_y + pool_h, ext_skimmer, ext_inferior)
    draw_corner_block(pool_x + pool_w, pool_y + pool_h, ext_escalera, ext_inferior)

    # Piscina
    draw.rounded_rectangle((pool_x, pool_y, pool_x + pool_w, pool_y + pool_h), radius=24, fill='#38bdf8', outline='#0369a1', width=6)
    draw.text((pool_x + 26, pool_y + 24), f"{pool_length:.2f} m x {pool_width:.2f} m", font=section_font, fill='#082f49')

    # Cotas legibles
    cota_y = pool_y - 84
    draw.line((pool_x, cota_y, pool_x + pool_w, cota_y), fill='#334155', width=4)
    draw.line((pool_x, cota_y - 16, pool_x, cota_y + 16), fill='#334155', width=4)
    draw.line((pool_x + pool_w, cota_y - 16, pool_x + pool_w, cota_y + 16), fill='#334155', width=4)
    draw.text((pool_x + pool_w / 2 - 110, cota_y - 54), f"LARGO {pool_length:.2f} m", font=section_font, fill='#0f172a')

    cota_x = pool_x - 120
    draw.line((cota_x, pool_y, cota_x, pool_y + pool_h), fill='#334155', width=4)
    draw.line((cota_x - 16, pool_y, cota_x + 16, pool_y), fill='#334155', width=4)
    draw.line((cota_x - 16, pool_y + pool_h, cota_x + 16, pool_y + pool_h), fill='#334155', width=4)
    draw.text((cota_x - 210, pool_y + pool_h / 2 - 14), f"ANCHO {pool_width:.2f} m", font=section_font, fill='#0f172a')

    if ext_skimmer > 0:
        draw.text((pool_x - ext_skimmer * tile_px + 18, pool_y + pool_h / 2 - 16), "SKIMMER", font=section_font, fill='#334155')
    if ext_escalera > 0:
        draw.text((pool_x + pool_w + 18, pool_y + pool_h / 2 - 16), "ESCALERA", font=section_font, fill='#334155')
    if ext_superior > 0:
        draw.text((pool_x + pool_w / 2 - 150, pool_y - ext_superior * tile_px - 54), "CABECERA SUPERIOR", font=section_font, fill='#334155')
    if ext_inferior > 0:
        draw.text((pool_x + pool_w / 2 - 150, pool_y + pool_h + ext_inferior * tile_px + 16), "CABECERA INFERIOR", font=section_font, fill='#334155')

    # Panel derecho
    panel_x = drawing_x0 + drawing_w + 30
    draw.rounded_rectangle((panel_x, drawing_y0, panel_x + info_panel_w, drawing_y0 + drawing_h), radius=20, fill='#0f172a')
    info_y = drawing_y0 + 30
    info_lines = [
        f"Cliente: {project_data.get('clientName', '-')}",
        f"Piscina: {pool.get('name', '-')}",
        f"Profundidad: {safe_float(pool.get('shallowDepth')):.2f} a {safe_float(pool.get('deepDepth')):.2f} m",
        "",
        "Capas por lado",
        f"Cabecera superior: {ext_superior}",
        f"Cabecera inferior: {ext_inferior}",
        f"Skimmer / izquierda: {ext_skimmer}",
        f"Escalera / derecha: {ext_escalera}",
        "",
        "Primer anillo",
        f"Norte: {side_configs['north'].get('firstRingType') or '-'}",
        f"Sur: {side_configs['south'].get('firstRingType') or '-'}",
        f"Este: {side_configs['east'].get('firstRingType') or '-'}",
        f"Oeste: {side_configs['west'].get('firstRingType') or '-'}",
    ]
    for line in info_lines:
        color = '#f8fafc' if line and not line.startswith('Capas') and not line.startswith('Primer') else '#93c5fd'
        draw.text((panel_x + 24, info_y), line, font=body_font, fill=color)
        info_y += 40 if line else 24

    fd, path = tempfile.mkstemp(suffix='.png', prefix='pool_tile_plan_')
    os.close(fd)
    img.save(path, format='PNG')
    return path

def add_tile_plan_sheet(wb, project_data, base_sheet_name):
    sheet_name = f"Plano {base_sheet_name}"[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Plano de Losetas'
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws['A2'] = f"Proyecto: {project_data.get('projectName', base_sheet_name)}"
    ws['A3'] = f"Cliente: {project_data.get('clientName', '-')}"
    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 18
    for row in range(1, 60):
        ws.row_dimensions[row].height = 22

    image_path = create_tile_plan_image(project_data)
    img = XLImage(image_path)
    img.width = 1400
    img.height = 930
    ws.add_image(img, 'A5')
    return image_path

def calculate_tile_summary(pool, tile_calculation, sidewalk):
    tile_calculation = tile_calculation or {}
    pool_length = safe_float(pool.get('length'))
    pool_width = safe_float(pool.get('width'))
    sidewalk_area = safe_float(sidewalk.get('materials', {}).get('area'))

    first_ring_type = None
    has_lomo = False
    first_ring_tiles = 0
    total_rows = 0
    extra_common_tiles = 0

    for side in ['north', 'south', 'east', 'west']:
        side_config = get_side_config(tile_calculation, side)
        side_length = pool_length if side in ['north', 'south'] else pool_width
        rows = safe_int(side_config.get('rows'))
        total_rows += rows

        if side_config.get('firstRingType'):
            first_ring_type = first_ring_type or side_config.get('firstRingType')
            first_ring_tiles += calculate_tiles_for_dimension(side_length)
            if normalize(side_config.get('firstRingType')) == 'lomo_ballena':
                has_lomo = True

        extra_rows = max(0, rows - (1 if side_config.get('firstRingType') else 0))
        if extra_rows > 0:
            extra_common_tiles += extra_rows * calculate_tiles_for_dimension(side_length)

    corner_tiles = 0
    north_rows = safe_int(get_side_config(tile_calculation, 'north').get('rows'))
    south_rows = safe_int(get_side_config(tile_calculation, 'south').get('rows'))
    east_rows = safe_int(get_side_config(tile_calculation, 'east').get('rows'))
    west_rows = safe_int(get_side_config(tile_calculation, 'west').get('rows'))

    if north_rows and west_rows:
        corner_tiles += north_rows * west_rows
    if north_rows and east_rows:
        corner_tiles += north_rows * east_rows
    if south_rows and west_rows:
        corner_tiles += south_rows * west_rows
    if south_rows and east_rows:
        corner_tiles += south_rows * east_rows
    if has_lomo:
        corner_tiles = max(0, corner_tiles - 4)

    total_common_tiles = extra_common_tiles + corner_tiles
    ring_area = sidewalk_area if sidewalk_area > 0 else ((pool_length + 0.96) * (pool_width + 0.96) - (pool_length * pool_width))

    return {
        'first_ring_type': normalize(first_ring_type or ''),
        'first_ring_label': 'Lomo Ballena' if has_lomo else 'Terminación',
        'first_ring_tiles': first_ring_tiles,
        'common_tiles': total_common_tiles,
        'corner_tiles': 4 if has_lomo else 0,
        'grate_tiles': safe_int(pool.get('canaletasCount', pool.get('canaletas', 0))),
        'sidewalk_area': round(sidewalk_area, 2),
        'first_ring_area': round(max(ring_area, 0), 2),
        'has_lomo': has_lomo,
        'total_rows': total_rows,
    }

def clear_visual_grid(ws):
    for row in range(11, 33):
        for col in range(16, 28):
            cell = ws.cell(row=row, column=col)
            cell.value = None

def ensure_grid_named_range(wb, ws):
    attr_text = f"'{ws.title}'!$P$11:$AA$32"
    wb.defined_names['grilla'] = DefinedName('grilla', attr_text=attr_text)

def draw_tile_grid(ws, pool, tile_calculation):
    clear_visual_grid(ws)

    top_row = 11
    bottom_row = 32
    left_col = 16   # P
    right_col = 27  # AA

    has_canaletas = safe_int(pool.get('canaletasCount', pool.get('canaletas', 0))) > 0
    common_marker = 'C'
    ring_marker = 'L'
    corner_marker = 'E'

    ring_type, _ = get_template_ring_type(tile_calculation or {})
    if ring_type != 'lomo':
        ring_marker = 'T'
        corner_marker = 'C'

    grid_height = bottom_row - top_row + 1  # 22
    grid_width = right_col - left_col + 1   # 12

    draw_width = grid_width - (1 if has_canaletas else 0)

    # Filas representan el largo de la piscina, columnas representan el ancho.
    pool_rows = max(1, min(grid_height - 2, safe_int(math.ceil(safe_float(pool.get('length')) / 0.5))))
    pool_cols = max(1, min(draw_width - 2, safe_int(math.ceil(safe_float(pool.get('width')) / 0.5))))

    side_north = get_side_config(tile_calculation, 'north')
    side_south = get_side_config(tile_calculation, 'south')
    side_east = get_side_config(tile_calculation, 'east')
    side_west = get_side_config(tile_calculation, 'west')

    # Como el largo se proyecta en filas, north/south se ubican en laterales verticales.
    left_layers = max(0, safe_int(side_north.get('rows')))
    right_layers = max(0, safe_int(side_south.get('rows')))
    top_layers = max(0, safe_int(side_east.get('rows')))
    bottom_layers = max(0, safe_int(side_west.get('rows')))

    total_height = pool_rows + top_layers + bottom_layers
    total_width = pool_cols + left_layers + right_layers

    if total_height > grid_height:
        overflow = total_height - grid_height
        trim_top = min(top_layers, math.ceil(overflow / 2))
        trim_bottom = min(bottom_layers, overflow - trim_top)
        top_layers -= trim_top
        bottom_layers -= trim_bottom
        total_height = pool_rows + top_layers + bottom_layers

    if total_width > draw_width:
        overflow = total_width - draw_width
        trim_left = min(left_layers, math.ceil(overflow / 2))
        trim_right = min(right_layers, overflow - trim_left)
        left_layers -= trim_left
        right_layers -= trim_right
        total_width = pool_cols + left_layers + right_layers

    top_padding = max(0, (grid_height - total_height) // 2)
    left_padding = max(0, (draw_width - total_width) // 2)

    pool_top = top_row + top_padding + top_layers
    pool_bottom = pool_top + pool_rows - 1
    pool_left = left_col + left_padding + left_layers
    pool_right = pool_left + pool_cols - 1

    ring_top = pool_top - 1 if top_layers > 0 else None
    ring_bottom = pool_bottom + 1 if bottom_layers > 0 else None
    ring_left = pool_left - 1 if left_layers > 0 else None
    ring_right = pool_right + 1 if right_layers > 0 else None

    # Dibuja filas/columnas externas de comunes si existen más de una capa hacia afuera.
    for layer in range(2, top_layers + 1):
        row = pool_top - layer
        for col in range(pool_left - left_layers, pool_right + right_layers + 1):
            ws.cell(row=row, column=col).value = common_marker
    for layer in range(2, bottom_layers + 1):
        row = pool_bottom + layer
        for col in range(pool_left - left_layers, pool_right + right_layers + 1):
            ws.cell(row=row, column=col).value = common_marker
    for layer in range(2, left_layers + 1):
        col = pool_left - layer
        for row in range(pool_top - top_layers, pool_bottom + bottom_layers + 1):
            ws.cell(row=row, column=col).value = common_marker
    for layer in range(2, right_layers + 1):
        col = pool_right + layer
        for row in range(pool_top - top_layers, pool_bottom + bottom_layers + 1):
            ws.cell(row=row, column=col).value = common_marker

    if ring_top is not None:
        for col in range(pool_left - (1 if left_layers > 0 else 0), pool_right + (1 if right_layers > 0 else 0) + 1):
            ws.cell(row=ring_top, column=col).value = ring_marker
    if ring_bottom is not None:
        for col in range(pool_left - (1 if left_layers > 0 else 0), pool_right + (1 if right_layers > 0 else 0) + 1):
            ws.cell(row=ring_bottom, column=col).value = ring_marker
    if ring_left is not None:
        for row in range(pool_top - (1 if top_layers > 0 else 0), pool_bottom + (1 if bottom_layers > 0 else 0) + 1):
            ws.cell(row=row, column=ring_left).value = ring_marker
    if ring_right is not None:
        for row in range(pool_top - (1 if top_layers > 0 else 0), pool_bottom + (1 if bottom_layers > 0 else 0) + 1):
            ws.cell(row=row, column=ring_right).value = ring_marker

    if ring_top is not None and ring_left is not None:
        ws.cell(row=ring_top, column=ring_left).value = corner_marker
    if ring_top is not None and ring_right is not None:
        ws.cell(row=ring_top, column=ring_right).value = corner_marker
    if ring_bottom is not None and ring_left is not None:
        ws.cell(row=ring_bottom, column=ring_left).value = corner_marker
    if ring_bottom is not None and ring_right is not None:
        ws.cell(row=ring_bottom, column=ring_right).value = corner_marker

    # Canaleta visual al lateral derecho.
    if has_canaletas:
        for row in range(top_row, bottom_row + 1):
            ws.cell(row=row, column=right_col).value = 'R'

def write_project_inputs(ws, project_data):
    pool = project_data.get('pool', {})
    support_bed = project_data.get('supportBed', {})
    plumbing = project_data.get('plumbing', {})
    electrical = project_data.get('electrical', {})

    support_materials = support_bed.get('materials', {}) or {}
    plumbing_items = plumbing.get('items', []) or []

    ws['B19'] = safe_int(support_materials.get('cement'))
    ws['B20'] = sum_plumbing_items(plumbing_items, keywords=['tee', 'te '])
    ws['B21'] = sum_plumbing_items(plumbing_items, keywords=['codo 90', '90'])
    ws['B22'] = sum_plumbing_items(plumbing_items, keywords=['codo 45', '45'])
    ws['B23'] = sum_plumbing_items(plumbing_items, keywords=['union simple', 'cupla simple'])
    ws['B24'] = sum_plumbing_items(plumbing_items, keywords=['union doble', 'cupla doble'])
    ws['B25'] = sum_plumbing_items(plumbing_items, keywords=['llave', 'valvula'])
    ws['A26'] = f"CAÑO {dominant_pipe_diameter(plumbing_items)}"
    ws['B26'] = sum_plumbing_items(plumbing_items, keywords=['caño', 'tubo'])
    ws['B27'] = max(1, round(sum_plumbing_items(plumbing_items, keywords=['pegamento', 'adhesivo', 'cemento pvc', 'solvente'])))
    ws['B28'] = round(safe_float(support_materials.get('geomembrane')))
    ws['B29'] = round(safe_float(electrical.get('distanceToPanel', 15)) * 2)
    ws['B30'] = round((safe_float(ws['B29'].value) / 1.5), 0) if ws['B29'].value else 0
    # B16 es una fórmula viva de la plantilla (mixto de la cama): NO sobrescribir.

def write_tile_outputs(ws, tile_summary, sidewalk):
    sidewalk_materials = sidewalk.get('materials', {}) or {}
    first_ring_count = tile_summary['first_ring_tiles']
    common_count = tile_summary['common_tiles']
    corner_count = tile_summary['corner_tiles']
    grate_count = tile_summary['grate_tiles']
    term_count = first_ring_count if tile_summary['first_ring_type'] != 'lomo_ballena' else 0
    lomo_count = first_ring_count if tile_summary['first_ring_type'] == 'lomo_ballena' else 0

    ws['E13'] = f"{first_ring_count} {tile_summary['first_ring_label']}" if first_ring_count else '-'
    ws['E14'] = corner_count if corner_count else '-'
    ws['E15'] = common_count if common_count else 0
    ws['E16'] = round(tile_summary['sidewalk_area'], 2)
    ws['E17'] = round(safe_float(sidewalk_materials.get('concreteVolume')), 2)
    ws['E18'] = safe_int(sidewalk_materials.get('cement'))
    ws['E19'] = safe_int(sidewalk_materials.get('sand'))
    ws['E20'] = safe_int(sidewalk_materials.get('stone'))
    ws['E21'] = safe_int(sidewalk_materials.get('mesh'))
    ws['E24'] = bags_from_kg(sidewalk_materials.get('adhesive'), 25)
    ws['E26'] = round((term_count * 0.24) + (lomo_count * 0.2), 2)
    ws['E27'] = round(safe_float(ws['K8'].value) * safe_float(ws['E26'].value), 2)
    ws['E28'] = round(max(safe_float(ws['K9'].value) - safe_float(ws['E27'].value), 0), 2)

    ws['H13'] = common_count if common_count else 0
    ws['I13'] = 0.25
    ws['J13'] = round(common_count * 0.25, 2)
    ws['H14'] = corner_count if corner_count else '-'
    ws['I14'] = 0.2
    ws['J14'] = round(corner_count * (0.4 * 0.4 - 0.05 * 0.05), 2) if corner_count else 0
    ws['H15'] = grate_count if grate_count else '-'
    ws['I15'] = 0.1
    ws['J15'] = round(grate_count * 0.1, 2) if grate_count else 0
    ws['H16'] = term_count if term_count else '-'
    ws['I16'] = 0.24
    ws['J16'] = round(term_count * 0.24, 2) if term_count else 0
    ws['H17'] = lomo_count if lomo_count else '-'
    ws['I17'] = 0.2
    ws['J17'] = round(lomo_count * 0.2, 2) if lomo_count else 0
    ws['J18'] = round(sum(safe_float(ws[cell].value) for cell in ['J13', 'J14', 'J15', 'J16', 'J17']), 2)

def normalize_template_formulas(ws):
    ws['E13'] = '=IF(LOWER($D$2)="terminacion",ROUNDUP(($A$2/$M$3*2)+($B$2/$M$2*2),0)&" Terminación",IF(LOWER($D$2)="lomo",ROUNDUP(($A$2/$M$3*2)+($B$2/$M$2*2)+4,0)&" Lomo Ballena","-"))'
    ws['E14'] = '=IF(UPPER($K$2)="LOMO",COUNTIF(grilla,"E"),"-")'
    ws['E15'] = '=COUNTIF(grilla,"C")'
    ws['H14'] = '=IF(COUNTIF(grilla,"E")=0,"-",COUNTIF(grilla,"E"))'
    ws['H15'] = '=IF(COUNTIF(grilla,"R")=0,"-",COUNTIF(grilla,"R"))'
    ws['H16'] = '=IF(COUNTIF(grilla,"T")=0,"-",COUNTIF(grilla,"T"))'
    ws['H17'] = '=IF(COUNTIF(grilla,"L")=0,"-",COUNTIF(grilla,"L"))'
    ws['J13'] = '=COUNTIF(grilla,"C")*0.25'
    ws['J14'] = '=COUNTIF(grilla,"E")*(0.4*0.4 - ($M$5/1000)^2)'
    ws['J15'] = '=COUNTIF(grilla,"R")*0.1'
    ws['J16'] = '=COUNTIF(grilla,"T")*0.24'
    ws['J17'] = '=COUNTIF(grilla,"L")*0.2'
    ws['J18'] = '=SUM($J$13:$J$17)'
    ws['B33'] = '=ROUND($B$15+$B$18+$E$19,0)&" m³"'
    ws['B34'] = '=ROUND($B$16+$E$26,0)&" m³"'

def get_template_ring_type(tile_calculation):
    first_ring_values = []
    for side in ['north', 'south', 'east', 'west']:
        side_data = tile_calculation.get(side, {}) or {}
        if side_data.get('firstRingType'):
            first_ring_values.append(normalize(side_data.get('firstRingType')))
    if any(value == 'lomo_ballena' for value in first_ring_values):
        return 'lomo', 'LOMO'
    return 'terminacion', 'TERMINACION'

def choose_template_sheet(workbook, pool):
    template_dimensions = {
        'CALCULOS DE MATERIALES TURQUESA': (6500, 3100),
        'CALCULO DE MATERIALES CIRCON': (5500, 2800),
        'CALCULOS MATERIALES GEMA AZUL': (7300, 3500),
    }

    available = [(name, dims) for name, dims in template_dimensions.items() if name in workbook.sheetnames]
    if not available:
        return workbook.sheetnames[0]

    target_length = int(round((pool.get('length', 0) or 0) * 1000))
    target_width = int(round((pool.get('width', 0) or 0) * 1000))

    def score(entry):
        _, (tmpl_length, tmpl_width) = entry
        return abs(tmpl_length - target_length) + abs(tmpl_width - target_width)

    return min(available, key=score)[0]

def add_image_to_cell(ws, image_url, cell_ref, width=100, height=100):
    """
    Agrega una imagen a una celda del Excel si existe

    Args:
        ws: Hoja de trabajo
        image_url: URL relativa de la imagen (ej: /uploads/products/equipment/bomba.jpg)
        cell_ref: Referencia de celda (ej: 'B10')
        width: Ancho de la imagen en píxeles
        height: Alto de la imagen en píxeles

    Returns:
        True si se agregó la imagen, False si no
    """
    try:
        # Construir ruta absoluta desde la URL relativa
        if not image_url or not image_url.startswith('/'):
            return False

        # Ruta base del proyecto
        base_path = Path(__file__).parent.parent  # backend/
        image_path = base_path / image_url.lstrip('/')

        if not image_path.exists():
            print(f"⚠ Imagen no encontrada: {image_path}")
            return False

        # Crear objeto de imagen
        img = XLImage(str(image_path))

        # Redimensionar
        img.width = width
        img.height = height

        # Agregar a la celda
        ws.add_image(img, cell_ref)
        print(f"✓ Imagen agregada: {image_path.name} en {cell_ref}")
        return True
    except Exception as e:
        print(f"Error al agregar imagen {image_url}: {e}")
        return False

def export_project_to_excel(excel_path, project_data):
    """
    Exporta un proyecto a una nueva hoja en el Excel siguiendo el formato existente

    Args:
        excel_path: Ruta al archivo Excel
        project_data: Diccionario con los datos del proyecto
    """

    temp_assets = []

    # Cargar el libro existente
    wb = openpyxl.load_workbook(excel_path)

    # Extraer datos anidados
    pool = project_data.get('pool', {})
    excavation = project_data.get('excavation', {})
    support_bed = project_data.get('supportBed', {})
    sidewalk = project_data.get('sidewalk', {})
    plumbing = project_data.get('plumbing', {})
    electrical = project_data.get('electrical', {})
    labor = project_data.get('labor', {})
    sections = project_data.get('sections', {})
    tile_calculation = project_data.get('tileCalculation', {}) or {}

    # Crear nombre de hoja basado en el proyecto
    sheet_name = f"{pool.get('name', 'Piscina')} - {project_data.get('clientName', 'Cliente')}"[:31]  # Max 31 caracteres

    # Si ya existe, eliminarla
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    template_name = choose_template_sheet(wb, pool)
    ws = wb.copy_worksheet(wb[template_name])
    ws.title = sheet_name

    first_ring_value, tile_type_value = get_template_ring_type(tile_calculation)
    tile_summary = calculate_tile_summary(pool, tile_calculation, sidewalk)
    layout_data = get_excel_layout_data(pool, tile_calculation)
    length_mm = int(round((pool.get('length', 0) or 0) * 1000))
    width_mm = int(round((pool.get('width', 0) or 0) * 1000))
    shallow_depth_mm = int(round((pool.get('shallowDepth', 0) or 0) * 1000))
    deep_depth_mm = int(round((pool.get('deepDepth', 0) or 0) * 1000))
    returns_count = int(pool.get('returnsCount', plumbing.get('returnsCount', 0)) or 0)

    ws['A2'] = length_mm
    ws['B2'] = width_mm
    ws['D2'] = first_ring_value
    ws['E2'] = shallow_depth_mm
    ws['F2'] = deep_depth_mm
    ws['G2'] = layout_data['stairs']
    ws['H2'] = layout_data['skimmer']
    ws['I2'] = layout_data['left_head']
    ws['J2'] = layout_data['right_head']
    ws['K2'] = tile_type_value
    ws['L2'] = 400 if tile_type_value == 'LOMO' else 480
    ws['M2'] = 500
    ws['N2'] = safe_int(pool.get('canaletasCount', pool.get('canaletas', 0)))

    ws['A12'] = project_data.get('responsible', 'Jesús Olguin')
    ws['D12'] = 'CHINO'

    ensure_grid_named_range(wb, ws)
    draw_tile_grid(ws, pool, tile_calculation)
    # La plantilla ya contiene TODAS las fórmulas vivas (volúmenes, áreas,
    # COUNTIF de la grilla, costos). Solo escribimos los INPUTS del proyecto y
    # dejamos que Excel recalcule. No se llama a normalize_template_formulas
    # (reescribía fórmulas con un bug $M$5/$M$6) ni a write_tile_outputs
    # (pisaba las fórmulas con valores estáticos), para que el Excel exportado
    # quede con fórmulas reales y recalcule al cambiar cualquier medida.
    write_project_inputs(ws, project_data)
    temp_assets.append(add_tile_plan_sheet(wb, project_data, sheet_name))

    for other_sheet in list(wb.sheetnames):
        if other_sheet not in {sheet_name, f"Plano {sheet_name}"[:31]}:
            del wb[other_sheet]

    wb.save(excel_path)
    for temp_path in temp_assets:
        try:
            os.unlink(temp_path)
        except OSError:
            pass
    print(f"✅ Hoja '{sheet_name}' creada desde plantilla '{template_name}'")
    return sheet_name

    # Crear nueva hoja
    ws = wb.create_sheet(sheet_name)

    # Configurar estilos
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    section_font = Font(name='Arial', size=12, bold=True)

    # Fila 2: Título
    ws['B2'] = 'Materiales e Instalación de Piscina'
    ws['B2'].font = title_font

    # Fila 4-8: Información del proyecto
    ws['B4'] = 'Fecha'
    ws['C4'] = datetime.now().strftime('%Y-%m-%d')
    ws['B5'] = 'Cliente'
    ws['C5'] = project_data.get('clientName', '-')
    ws['B6'] = 'Domicilio'
    ws['C6'] = project_data.get('address', '-')
    ws['B7'] = 'Piscina'
    ws['C7'] = f"{pool.get('name', '-')} ({pool.get('length', 0)}×{pool.get('width', 0)} m x {pool.get('shallowDepth', 0)} a {pool.get('deepDepth', 0)} metros de profundidad)"
    ws['B8'] = 'Volumen'
    ws['C8'] = f"{pool.get('volume', 0):.2f} m³ / Espejo de agua: {pool.get('waterMirrorArea', 0):.2f} m²"

    # Aplicar negrita a las etiquetas
    for row in [4, 5, 6, 7, 8]:
        ws[f'B{row}'].font = header_font

    # Fila 10: Encabezados de tabla
    ws['B10'] = 'Materiales / Consumible'
    ws['C10'] = 'Diámetro/Tipo'
    ws['D10'] = 'Unidad'
    ws['E10'] = 'Cantidad'
    ws['F10'] = 'Observaciones'
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}10'].font = header_font

    # Fila 12: Responsable
    ws['B12'] = project_data.get('responsible', 'Jesús Olguin')
    ws['B12'].font = header_font

    current_row = 14

    # ===== SECCIÓN: EXCAVACIÓN =====
    if sections.get('excavation', True):
        ws[f'B{current_row}'] = 'EXCAVACIÓN'
        ws[f'B{current_row}'].font = section_font
        current_row += 1

        ws[f'B{current_row}'] = 'Longitud de excavación'
        ws[f'D{current_row}'] = 'm'
        ws[f'E{current_row}'] = excavation.get('length', 0)
        current_row += 1

        ws[f'B{current_row}'] = 'Ancho de excavación'
        ws[f'D{current_row}'] = 'm'
        ws[f'E{current_row}'] = excavation.get('width', 0)
        current_row += 1

        ws[f'B{current_row}'] = 'Profundidad de excavación'
        ws[f'D{current_row}'] = 'm'
        ws[f'E{current_row}'] = excavation.get('depth', 0)
        current_row += 1

        ws[f'B{current_row}'] = 'Volumen total de excavación'
        ws[f'D{current_row}'] = 'm³'
        ws[f'E{current_row}'] = excavation.get('volume', 0)
        current_row += 2

    # ===== SECCIÓN: CAMA DE APOYO =====
    if sections.get('supportBed', True):
        ws[f'B{current_row}'] = 'CAMA DE APOYO'
        ws[f'B{current_row}'].font = section_font
        current_row += 1

        materials = support_bed.get('materials', {})
        ws[f'B{current_row}'] = 'Cemento para la cama'
        ws[f'D{current_row}'] = materials.get('cementUnit', 'bolsas')
        ws[f'E{current_row}'] = materials.get('cement', 0)
        current_row += 1

        ws[f'B{current_row}'] = 'Arena gruesa'
        ws[f'D{current_row}'] = materials.get('sandUnit', 'm³')
        ws[f'E{current_row}'] = materials.get('sand', 0)
        current_row += 1

        ws[f'B{current_row}'] = 'Mixto para la cama'
        ws[f'D{current_row}'] = materials.get('mixedUnit', 'm³')
        ws[f'E{current_row}'] = materials.get('mixed', 0)
        current_row += 2

    # ===== SECCIÓN: VEREDA =====
    if sections.get('sidewalk', True):
        ws[f'B{current_row}'] = 'VEREDA'
        ws[f'B{current_row}'].font = section_font
        current_row += 1

        materials = sidewalk.get('materials', {})
        ws[f'B{current_row}'] = 'Cemento para vereda'
        ws[f'D{current_row}'] = materials.get('cementUnit', 'bolsas')
        ws[f'E{current_row}'] = materials.get('cement', 0)
        current_row += 1

        ws[f'B{current_row}'] = 'Arena para vereda'
        ws[f'D{current_row}'] = materials.get('sandUnit', 'm³')
        ws[f'E{current_row}'] = materials.get('sand', 0)
        current_row += 1

        ws[f'B{current_row}'] = 'Piedra para vereda'
        ws[f'D{current_row}'] = materials.get('stoneUnit', 'm³')
        ws[f'E{current_row}'] = materials.get('stone', 0)
        current_row += 1

        ws[f'B{current_row}'] = 'Malla sima'
        ws[f'D{current_row}'] = materials.get('meshUnit', 'unidad')
        ws[f'E{current_row}'] = materials.get('mesh', 0)
        current_row += 2

    # ===== SECCIÓN: PLOMERÍA =====
    if sections.get('plumbing', True):
        ws[f'B{current_row}'] = 'PLOMERÍA Y MATERIALES PVC'
        ws[f'B{current_row}'].font = section_font
        current_row += 1

        # Items de plomería con detalles completos
        plumbing_items = plumbing.get('items', [])
        if plumbing_items:
            for item in plumbing_items:
                ws[f'B{current_row}'] = item.get('name', '-')
                ws[f'C{current_row}'] = item.get('diameter', '-')
                ws[f'D{current_row}'] = item.get('type', 'PVC')
                ws[f'E{current_row}'] = item.get('quantity', 0)
                ws[f'F{current_row}'] = item.get('observations', '-')
                current_row += 1
        else:
            ws[f'B{current_row}'] = 'Sin items de plomería especificados'
            current_row += 1

        current_row += 1

    # ===== SECCIÓN: ELÉCTRICA =====
    if sections.get('electrical', True):
        ws[f'B{current_row}'] = 'INSTALACIÓN ELÉCTRICA Y EQUIPOS'
        ws[f'B{current_row}'].font = section_font
        current_row += 1

        # Bomba (con imagen si está disponible)
        pump = electrical.get('pump', {})
        pump_image_url = pump.get('imageUrl', None)

        ws[f'B{current_row}'] = 'Bomba'
        ws[f'C{current_row}'] = pump.get('power', '-')
        ws[f'E{current_row}'] = 1
        ws[f'F{current_row}'] = pump.get('observations', '-')

        # Intentar agregar imagen de la bomba
        if pump_image_url:
            # La imagen se inserta en la columna A de las siguientes filas
            image_added = add_image_to_cell(ws, pump_image_url, f'A{current_row}', width=80, height=80)
            if image_added:
                ws.row_dimensions[current_row].height = 60  # Ajustar altura de fila

        current_row += 1

        # Filtro (con imagen si está disponible)
        filter_data = electrical.get('filter', {})
        filter_image_url = filter_data.get('imageUrl', None)

        ws[f'B{current_row}'] = 'Filtro'
        ws[f'C{current_row}'] = filter_data.get('diameter', '-')
        ws[f'E{current_row}'] = 1
        ws[f'F{current_row}'] = filter_data.get('observations', '-')

        # Intentar agregar imagen del filtro
        if filter_image_url:
            image_added = add_image_to_cell(ws, filter_image_url, f'A{current_row}', width=80, height=80)
            if image_added:
                ws.row_dimensions[current_row].height = 60  # Ajustar altura de fila

        current_row += 1

        # Especificaciones eléctricas
        ws[f'B{current_row}'] = 'Consumo total'
        ws[f'E{current_row}'] = f"{electrical.get('watts', 0)} W"
        current_row += 1

        ws[f'B{current_row}'] = 'Amperaje'
        ws[f'E{current_row}'] = f"{electrical.get('amps', 0)} A"
        current_row += 1

        # Consumo desglosado
        consumption = electrical.get('consumptionBreakdown', [])
        if consumption:
            ws[f'B{current_row}'] = 'Desglose de consumo:'
            ws[f'B{current_row}'].font = header_font
            current_row += 1
            for item in consumption:
                ws[f'B{current_row}'] = f"  {item.get('item', '-')}"
                ws[f'E{current_row}'] = f"{item.get('watts', 0)} W"
                current_row += 1

        current_row += 1

    # ===== SECCIÓN: ANÁLISIS HIDRÁULICO PROFESIONAL =====
    hydraulic_analysis = project_data.get('hydraulicAnalysis', None)
    if sections.get('hydraulicAnalysis', True) and hydraulic_analysis:
        ws[f'B{current_row}'] = 'ANÁLISIS HIDRÁULICO PROFESIONAL'
        ws[f'B{current_row}'].font = section_font
        current_row += 1

        # TDH (Total Dynamic Head)
        tdh = hydraulic_analysis.get('totalDynamicHead', 0)
        ws[f'B{current_row}'] = 'TDH Total (Altura Dinámica Total)'
        ws[f'E{current_row}'] = f"{tdh:.2f} m"
        ws[f'F{current_row}'] = 'Altura que debe vencer la bomba'
        current_row += 1

        # Pérdidas por fricción
        friction = hydraulic_analysis.get('frictionLoss', {})
        ws[f'B{current_row}'] = 'Pérdida por fricción (succión)'
        ws[f'E{current_row}'] = f"{friction.get('suction', 0):.2f} m"
        current_row += 1

        ws[f'B{current_row}'] = 'Pérdida por fricción (retorno)'
        ws[f'E{current_row}'] = f"{friction.get('return', 0):.2f} m"
        current_row += 1

        ws[f'B{current_row}'] = 'Pérdida por fricción total'
        ws[f'E{current_row}'] = f"{friction.get('total', 0):.2f} m"
        ws[f'B{current_row}'].font = header_font
        current_row += 1

        # Pérdidas singulares
        singular = hydraulic_analysis.get('singularLoss', {})
        ws[f'B{current_row}'] = 'Pérdida singular (accesorios succión)'
        ws[f'E{current_row}'] = f"{singular.get('suction', 0):.2f} m"
        current_row += 1

        ws[f'B{current_row}'] = 'Pérdida singular (accesorios retorno)'
        ws[f'E{current_row}'] = f"{singular.get('return', 0):.2f} m"
        current_row += 1

        ws[f'B{current_row}'] = 'Pérdida singular total'
        ws[f'E{current_row}'] = f"{singular.get('total', 0):.2f} m"
        ws[f'B{current_row}'].font = header_font
        current_row += 1

        # Validaciones de velocidad
        velocity_checks = hydraulic_analysis.get('velocityChecks', [])
        if velocity_checks:
            ws[f'B{current_row}'] = 'Validación de velocidades:'
            ws[f'B{current_row}'].font = header_font
            current_row += 1
            for check in velocity_checks:
                line_type = check.get('lineType', '-')
                velocity = check.get('velocity', 0)
                is_ok = check.get('isOk', False)
                status = '✓ OK' if is_ok else '⚠ Fuera de rango'
                ws[f'B{current_row}'] = f"  {line_type}: {velocity:.2f} m/s - {status}"
                ws[f'F{current_row}'] = 'Rango óptimo: 1.5-2.5 m/s'
                current_row += 1

        # Advertencias
        warnings = hydraulic_analysis.get('warnings', [])
        if warnings:
            ws[f'B{current_row}'] = 'Advertencias:'
            ws[f'B{current_row}'].font = header_font
            current_row += 1
            for warning in warnings:
                ws[f'B{current_row}'] = f"  ⚠ {warning}"
                current_row += 1

        # Bomba recomendada
        recommended_pump = hydraulic_analysis.get('recommendedPump', None)
        if recommended_pump:
            ws[f'B{current_row}'] = 'Bomba seleccionada según TDH:'
            ws[f'B{current_row}'].font = header_font
            current_row += 1

            pump_name = recommended_pump.get('name', '-')
            pump_flow_rate = recommended_pump.get('flowRate', '-')
            pump_description = recommended_pump.get('description', '-')
            pump_image_url = recommended_pump.get('imageUrl', None)

            ws[f'B{current_row}'] = f"  {pump_name}"
            ws[f'C{current_row}'] = pump_flow_rate
            ws[f'F{current_row}'] = pump_description

            # Intentar agregar imagen de la bomba recomendada
            if pump_image_url:
                image_added = add_image_to_cell(ws, pump_image_url, f'A{current_row}', width=80, height=80)
                if image_added:
                    ws.row_dimensions[current_row].height = 60  # Ajustar altura de fila

            current_row += 1

        current_row += 1

    # ===== SECCIÓN: ANÁLISIS ELÉCTRICO PROFESIONAL =====
    electrical_analysis = project_data.get('electricalAnalysis', None)
    if sections.get('electricalAnalysis', True) and electrical_analysis:
        ws[f'B{current_row}'] = 'ANÁLISIS ELÉCTRICO PROFESIONAL'
        ws[f'B{current_row}'].font = section_font
        current_row += 1

        # Potencia total
        total_power = electrical_analysis.get('totalPowerInstalled', 0)
        demand_power = electrical_analysis.get('totalPowerDemand', 0)
        total_current = electrical_analysis.get('totalCurrent', 0)

        ws[f'B{current_row}'] = 'Potencia instalada total'
        ws[f'E{current_row}'] = f"{total_power:.0f} W"
        current_row += 1

        ws[f'B{current_row}'] = 'Potencia de demanda (con simultaneidad)'
        ws[f'E{current_row}'] = f"{demand_power:.0f} W"
        ws[f'F{current_row}'] = 'Considera factor de simultaneidad'
        current_row += 1

        ws[f'B{current_row}'] = 'Corriente total calculada'
        ws[f'E{current_row}'] = f"{total_current:.2f} A"
        ws[f'F{current_row}'] = 'Con factor de potencia y eficiencia'
        ws[f'B{current_row}'].font = header_font
        current_row += 1

        # Cable dimensionado
        cable = electrical_analysis.get('cable', {})
        ws[f'B{current_row}'] = 'Cable recomendado'
        ws[f'C{current_row}'] = f"{cable.get('section', '-')} mm²"
        ws[f'F{current_row}'] = f"Caída de tensión: {cable.get('voltageDrop', 0):.2f}% (máx 3%)"
        current_row += 1

        ws[f'B{current_row}'] = 'Capacidad de corriente del cable'
        ws[f'E{current_row}'] = f"{cable.get('currentCapacity', 0):.1f} A"
        current_row += 1

        # Protecciones
        protection = electrical_analysis.get('protection', {})
        ws[f'B{current_row}'] = 'Interruptor termomagnético (Breaker)'
        ws[f'C{current_row}'] = f"{protection.get('breakerSize', 0)} A"
        ws[f'F{current_row}'] = 'Curva C recomendada'
        current_row += 1

        ws[f'B{current_row}'] = 'Diferencial (RCD)'
        ws[f'C{current_row}'] = f"{protection.get('rcdSize', 0)} A / 30 mA"
        ws[f'F{current_row}'] = 'Obligatorio para piscinas'
        current_row += 1

        # Cargas individuales
        loads = electrical_analysis.get('loads', [])
        if loads:
            ws[f'B{current_row}'] = 'Desglose de cargas eléctricas:'
            ws[f'B{current_row}'].font = header_font
            current_row += 1
            ws[f'B{current_row}'] = 'Equipo'
            ws[f'C{current_row}'] = 'Potencia (W)'
            ws[f'D{current_row}'] = 'Corriente (A)'
            ws[f'E{current_row}'] = 'FP (cos φ)'
            ws[f'F{current_row}'] = 'Observaciones'
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{current_row}'].font = header_font
            current_row += 1

            for load in loads:
                ws[f'B{current_row}'] = load.get('name', '-')
                ws[f'C{current_row}'] = f"{load.get('power', 0)} W"
                ws[f'D{current_row}'] = f"{load.get('current', 0):.2f} A"
                ws[f'E{current_row}'] = f"{load.get('powerFactor', 1):.2f}"
                ws[f'F{current_row}'] = f"Eficiencia: {load.get('efficiency', 1):.0%}"
                current_row += 1

        # Costo operativo
        operating_cost = electrical_analysis.get('operatingCost', {})
        if operating_cost:
            ws[f'B{current_row}'] = 'Costo operativo estimado:'
            ws[f'B{current_row}'].font = header_font
            current_row += 1

            ws[f'B{current_row}'] = 'Consumo diario'
            ws[f'E{current_row}'] = f"{operating_cost.get('dailyKwh', 0):.2f} kWh"
            current_row += 1

            ws[f'B{current_row}'] = 'Costo mensual estimado'
            ws[f'E{current_row}'] = f"${operating_cost.get('monthlyCost', 0):.2f}"
            ws[f'F{current_row}'] = '8 hrs/día promedio'
            current_row += 1

            ws[f'B{current_row}'] = 'Costo anual estimado'
            ws[f'E{current_row}'] = f"${operating_cost.get('annualCost', 0):.2f}"
            current_row += 1

        # Advertencias eléctricas
        elec_warnings = electrical_analysis.get('warnings', [])
        if elec_warnings:
            ws[f'B{current_row}'] = 'Advertencias eléctricas:'
            ws[f'B{current_row}'].font = header_font
            current_row += 1
            for warning in elec_warnings:
                ws[f'B{current_row}'] = f"  ⚠ {warning}"
                current_row += 1

        current_row += 1

    # ===== SECCIÓN: MANO DE OBRA =====
    if sections.get('labor', True):
        ws[f'B{current_row}'] = 'MANO DE OBRA'
        ws[f'B{current_row}'].font = section_font
        current_row += 1

        roles = labor.get('roles', [])
        if roles:
            ws[f'B{current_row}'] = 'Rol'
            ws[f'C{current_row}'] = 'Tareas'
            ws[f'D{current_row}'] = 'Horas'
            ws[f'E{current_row}'] = 'Costo'
            for col in ['B', 'C', 'D', 'E']:
                ws[f'{col}{current_row}'].font = header_font
            current_row += 1

            total_labor_cost = 0
            for role in roles:
                ws[f'B{current_row}'] = role.get('role', '-')
                ws[f'C{current_row}'] = role.get('tasks', 0)
                ws[f'D{current_row}'] = role.get('hours', 0)
                ws[f'E{current_row}'] = f"${role.get('cost', 0):,.2f}"
                total_labor_cost += role.get('cost', 0)
                current_row += 1

            # Total de mano de obra
            ws[f'B{current_row}'] = 'TOTAL MANO DE OBRA'
            ws[f'B{current_row}'].font = header_font
            ws[f'E{current_row}'] = f"${total_labor_cost:,.2f}"
            ws[f'E{current_row}'].font = header_font
            current_row += 1
        else:
            ws[f'B{current_row}'] = 'Sin mano de obra especificada'
            current_row += 1

        current_row += 1

    # ===== SECCIÓN: SECUENCIA DE TRABAJO =====
    if sections.get('sequence', True):
        ws[f'B{current_row}'] = 'SECUENCIA DE TRABAJO'
        ws[f'B{current_row}'].font = section_font
        current_row += 1

        ws[f'B{current_row}'] = '1. Marcado y replanteo del terreno'
        current_row += 1
        ws[f'B{current_row}'] = '2. Excavación según dimensiones especificadas'
        current_row += 1
        ws[f'B{current_row}'] = '3. Preparación de cama de apoyo'
        current_row += 1
        ws[f'B{current_row}'] = '4. Instalación de la piscina'
        current_row += 1
        ws[f'B{current_row}'] = '5. Instalación hidráulica (plomería)'
        current_row += 1
        ws[f'B{current_row}'] = '6. Instalación eléctrica y equipos'
        current_row += 1
        ws[f'B{current_row}'] = '7. Relleno y compactación'
        current_row += 1
        ws[f'B{current_row}'] = '8. Construcción de vereda perimetral'
        current_row += 1
        ws[f'B{current_row}'] = '9. Pruebas de funcionamiento'
        current_row += 2

    # ===== SECCIÓN: NORMAS Y OBSERVACIONES =====
    if sections.get('standards', True):
        ws[f'B{current_row}'] = 'NORMAS Y OBSERVACIONES'
        ws[f'B{current_row}'].font = section_font
        current_row += 1

        ws[f'B{current_row}'] = '• Todos los materiales deben cumplir con normas IRAM vigentes'
        current_row += 1
        ws[f'B{current_row}'] = '• La instalación eléctrica debe ser realizada por electricista matriculado'
        current_row += 1
        ws[f'B{current_row}'] = '• Los equipos deben instalarse en lugar ventilado y protegido'
        current_row += 1
        ws[f'B{current_row}'] = '• Realizar prueba de estanqueidad antes del llenado final'
        current_row += 1
        ws[f'B{current_row}'] = '• Coordinar con albañil para trabajos de vereda y relleno'
        current_row += 1

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30

    # Guardar el archivo
    wb.save(excel_path)
    print(f"✅ Hoja '{sheet_name}' agregada exitosamente al Excel")
    return sheet_name

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python export_to_excel.py <json_data> [excel_path]")
        sys.exit(1)

    # Leer datos del proyecto desde JSON
    project_data = json.loads(sys.argv[1])

    # Ruta al archivo Excel
    excel_path = sys.argv[2] if len(sys.argv) > 2 else '/home/jesusolguin/Projects/pool-calculator/backend/public/CALCULADORA MATERIALES AQUAM.xlsx'

    # Exportar
    export_project_to_excel(excel_path, project_data)
