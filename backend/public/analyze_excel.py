#!/usr/bin/env python3
import openpyxl
import json
from collections import defaultdict

# Cargar el archivo Excel
wb = openpyxl.load_workbook('CALCULADORA MATERIALES AQUAM.xlsx', data_only=True)

# Listar todas las hojas
print("=" * 80)
print("HOJAS DISPONIBLES EN EL ARCHIVO:")
print("=" * 80)
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"{idx}. {sheet_name}")

print("\n" + "=" * 80)
print("BUSCANDO HOJAS RELACIONADAS CON 'TURQUESA' E 'INÉS':")
print("=" * 80)

# Buscar hojas relevantes
turquesa_sheets = []
for sheet_name in wb.sheetnames:
    if 'turquesa' in sheet_name.lower() or 'ines' in sheet_name.lower() or 'inés' in sheet_name.lower():
        turquesa_sheets.append(sheet_name)
        print(f"✓ Encontrada: {sheet_name}")

if not turquesa_sheets:
    print("No se encontraron hojas con 'Turquesa' o 'Inés' en el nombre")
else:
    print(f"\nTotal de hojas relevantes: {len(turquesa_sheets)}")

# Función para extraer datos de una hoja
def extract_sheet_data(sheet_name):
    sheet = wb[sheet_name]
    data = {}

    print(f"\n{'=' * 80}")
    print(f"ANÁLISIS DE HOJA: {sheet_name}")
    print(f"{'=' * 80}")

    # Extraer dimensiones
    print(f"Dimensiones: {sheet.max_row} filas x {sheet.max_column} columnas")

    # Extraer todos los valores no vacíos
    values = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(50, sheet.max_row), values_only=True), 1):
        row_data = []
        for col_idx, cell in enumerate(row, 1):
            if cell is not None and str(cell).strip():
                row_data.append({
                    'col': col_idx,
                    'value': cell
                })
        if row_data:
            values.append({
                'row': row_idx,
                'data': row_data
            })

    # Mostrar primeras filas con datos
    print(f"\nPrimeras filas con datos (hasta 30):")
    for item in values[:30]:
        row_num = item['row']
        row_str = f"Fila {row_num}: "
        cells = []
        for cell_data in item['data']:
            cells.append(f"Col{cell_data['col']}={cell_data['value']}")
        print(row_str + " | ".join(cells))

    return values

# Extraer datos de las hojas relevantes
sheets_data = {}
for sheet_name in turquesa_sheets:
    sheets_data[sheet_name] = extract_sheet_data(sheet_name)

# Si hay exactamente 2 hojas, hacer comparación
if len(turquesa_sheets) == 2:
    print(f"\n{'=' * 80}")
    print(f"COMPARACIÓN ENTRE: {turquesa_sheets[0]} VS {turquesa_sheets[1]}")
    print(f"{'=' * 80}")

    sheet1_data = sheets_data[turquesa_sheets[0]]
    sheet2_data = sheets_data[turquesa_sheets[1]]

    print(f"\n{turquesa_sheets[0]}: {len(sheet1_data)} filas con datos")
    print(f"{turquesa_sheets[1]}: {len(sheet2_data)} filas con datos")

print("\n" + "=" * 80)
print("ANÁLISIS COMPLETADO")
print("=" * 80)
