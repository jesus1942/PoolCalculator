#!/usr/bin/env python3
import openpyxl
from openpyxl.utils import get_column_letter

# Cargar el archivo Excel
wb = openpyxl.load_workbook('CALCULADORA MATERIALES AQUAM.xlsx', data_only=True)

# Hojas a comparar
sheet1_name = 'CALCULOS DE MATERIALES TURQUESA'
sheet2_name = 'TURQUESA (6,5X3,1) INES Y PABLO'

sheet1 = wb[sheet1_name]
sheet2 = wb[sheet2_name]

print("=" * 100)
print(f"COMPARACIÓN DETALLADA: PISCINA TURQUESA")
print("=" * 100)

# PARTE 1: Extraer datos principales de la hoja de cálculos
print("\n" + "=" * 100)
print(f"HOJA 1: {sheet1_name}")
print("=" * 100)

# Dimensiones de la piscina
print("\n📏 DIMENSIONES DE LA PISCINA:")
largo = sheet1.cell(2, 1).value  # A2
ancho = sheet1.cell(2, 2).value  # B2
pando = sheet1.cell(2, 5).value  # E2
hondo = sheet1.cell(2, 6).value  # F2

print(f"  Largo: {largo} mm = {largo/1000 if largo else 'N/A'} m")
print(f"  Ancho: {ancho} mm = {ancho/1000 if ancho else 'N/A'} m")
print(f"  Parte poco profunda: {pando} mm = {pando/1000 if pando else 'N/A'} m")
print(f"  Parte profunda: {hondo} mm = {hondo/1000 if hondo else 'N/A'} m")

# Materiales de PVC
print("\n🔧 MATERIALES DE PVC (40mm) - HOJA DE CÁLCULOS:")
materiales_pvc_calc = {
    'TEE': sheet1.cell(20, 2).value,
    'CODO 90°': sheet1.cell(21, 2).value,
    'CODO 45°': sheet1.cell(22, 2).value,
    'UNION SIMPLE': sheet1.cell(23, 2).value,
    'UNION DOBLE': sheet1.cell(24, 2).value,
    'LLAVE ESFERICA': sheet1.cell(25, 2).value,
    'CAÑO 40MM': sheet1.cell(26, 2).value,
}

for material, cantidad in materiales_pvc_calc.items():
    print(f"  {material}: {cantidad}")

# Otros materiales importantes
print("\n📦 OTROS MATERIALES - HOJA DE CÁLCULOS:")
otros_calc = {
    'PEGAMENTO x 1kg': sheet1.cell(27, 2).value,
    'GEOTEXTIL O GEOMEMBRANA 400MICRONES X METRO': sheet1.cell(28, 2).value,
    'CAÑO AZUL 25MM X METRO': sheet1.cell(29, 2).value,
    'CONECTORES PARA CORRUGADO DE 25MM': sheet1.cell(30, 2).value,
}

for material, cantidad in otros_calc.items():
    print(f"  {material}: {cantidad}")

# Materiales de construcción
print("\n🏗️ MATERIALES DE CONSTRUCCIÓN - HOJA DE CÁLCULOS:")
construccion_calc = {
    'SUPERFICIE DE LA CAMA': sheet1.cell(13, 2).value,
    'VOLUMEN DE LA CAMA M3': sheet1.cell(14, 2).value,
    'ARENA M3': sheet1.cell(15, 2).value,
    'MIXTO M3': sheet1.cell(16, 2).value,
    'MALLA': sheet1.cell(17, 2).value,
    'ARENA PARA TAPADO M3': sheet1.cell(18, 2).value,
    'CEMENTO': sheet1.cell(19, 2).value,
}

for material, cantidad in construccion_calc.items():
    print(f"  {material}: {cantidad}")

# Cálculos de vereda y losetas
print("\n🔲 CÁLCULOS DE VEREDA Y LOSETAS:")
vereda_calc = {
    'PRIMERA HILADA O ANILLO': sheet1.cell(13, 5).value,
    'LOSETAS COMUNES': sheet1.cell(15, 5).value,
    'M2 DE VEREDA COMPLETA': sheet1.cell(16, 5).value,
    'M3 CONTRAPISO VEREDA': sheet1.cell(17, 5).value,
    'BOLSAS DE CEMENTO': sheet1.cell(18, 5).value,
    'ARENA M3 (vereda)': sheet1.cell(19, 5).value,
    'PIEDRA': sheet1.cell(20, 5).value,
    'MALLA (vereda)': sheet1.cell(21, 5).value,
    'BOLSAS DE KLAUKOL': sheet1.cell(24, 5).value,
}

for item, valor in vereda_calc.items():
    print(f"  {item}: {valor}")

# PARTE 2: Extraer datos de la hoja del proyecto de Inés y Pablo
print("\n" + "=" * 100)
print(f"HOJA 2: {sheet2_name}")
print("=" * 100)

# Información del cliente
print("\n👥 INFORMACIÓN DEL CLIENTE:")
fecha = sheet2.cell(4, 3).value
cliente = sheet2.cell(5, 3).value
domicilio = sheet2.cell(6, 3).value
piscina = sheet2.cell(7, 3).value

print(f"  Fecha: {fecha}")
print(f"  Cliente: {cliente}")
print(f"  Domicilio: {domicilio}")
print(f"  Piscina: {piscina}")

# Materiales de PVC del proyecto
print("\n🔧 MATERIALES DE PVC (40mm) - PROYECTO INÉS Y PABLO:")
materiales_pvc_proyecto = {
    'Codo 90°': sheet2.cell(14, 5).value,
    'Codo 45°': sheet2.cell(15, 5).value,
    'Tee': sheet2.cell(16, 5).value,
    'Válvula de corte (esférica)': sheet2.cell(17, 5).value,
    'Caños PN10 x 6 m': sheet2.cell(18, 5).value,
}

for material, cantidad in materiales_pvc_proyecto.items():
    print(f"  {material}: {cantidad}")

# Adicionales del proyecto
print("\n📦 ADICIONALES - PROYECTO INÉS Y PABLO:")
adicionales_proyecto = {
    'Pegamento PVC azul (Tigre) x 500 ml': sheet2.cell(22, 5).value,
    'Acetona x 1 Litro': sheet2.cell(23, 5).value,
    'Cinta teflón 1 rollo ALTA DENSIDAD': sheet2.cell(24, 5).value,
    'Pasta selladora 1 tubo': sheet2.cell(25, 5).value,
}

for material, cantidad in adicionales_proyecto.items():
    print(f"  {material}: {cantidad}")

# Materiales de construcción del proyecto
print("\n🏗️ MATERIALES DE CONSTRUCCIÓN - PROYECTO INÉS Y PABLO:")
construccion_proyecto = {
    'Malla sima 6 mm': sheet2.cell(35, 5).value,
    'Mixto para relleno (m3)': sheet2.cell(37, 5).value,
    'Mixto para la cama (m3)': sheet2.cell(38, 5).value,
    'Arena Gruesa m3': sheet2.cell(39, 5).value,
}

for material, cantidad in construccion_proyecto.items():
    print(f"  {material}: {cantidad}")

# PARTE 3: COMPARACIÓN Y ANÁLISIS DE DISCREPANCIAS
print("\n" + "=" * 100)
print("⚠️  ANÁLISIS DE DISCREPANCIAS Y DIFERENCIAS")
print("=" * 100)

print("\n🔍 COMPARACIÓN DE MATERIALES DE PVC (40mm):")
comparaciones_pvc = [
    ('TEE', materiales_pvc_calc.get('TEE'), 'Tee', materiales_pvc_proyecto.get('Tee')),
    ('CODO 90°', materiales_pvc_calc.get('CODO 90°'), 'Codo 90°', materiales_pvc_proyecto.get('Codo 90°')),
    ('CODO 45°', materiales_pvc_calc.get('CODO 45°'), 'Codo 45°', materiales_pvc_proyecto.get('Codo 45°')),
    ('LLAVE ESFERICA', materiales_pvc_calc.get('LLAVE ESFERICA'), 'Válvula de corte (esférica)', materiales_pvc_proyecto.get('Válvula de corte (esférica)')),
    ('CAÑO 40MM', materiales_pvc_calc.get('CAÑO 40MM'), 'Caños PN10 x 6 m', materiales_pvc_proyecto.get('Caños PN10 x 6 m')),
]

discrepancias_encontradas = False
for nombre_calc, cant_calc, nombre_proy, cant_proy in comparaciones_pvc:
    if cant_calc != cant_proy:
        discrepancias_encontradas = True
        diferencia = cant_proy - cant_calc if (cant_proy and cant_calc) else 'N/A'
        print(f"  ❌ {nombre_calc}:")
        print(f"     Hoja de cálculos: {cant_calc}")
        print(f"     Proyecto: {cant_proy}")
        print(f"     Diferencia: {diferencia}")
    else:
        print(f"  ✅ {nombre_calc}: {cant_calc} (coinciden)")

print("\n🔍 COMPARACIÓN DE MATERIALES DE CONSTRUCCIÓN:")

# Comparar mixto
mixto_calc = construccion_calc.get('MIXTO M3')
mixto_cama_proy = construccion_proyecto.get('Mixto para la cama (m3)')
mixto_relleno_proy = construccion_proyecto.get('Mixto para relleno (m3)')

print(f"  MIXTO:")
print(f"     Hoja de cálculos (cama): {mixto_calc} m³")
print(f"     Proyecto (cama): {mixto_cama_proy} m³")
print(f"     Proyecto (relleno): {mixto_relleno_proy} m³")
if mixto_calc and mixto_cama_proy:
    if mixto_calc != mixto_cama_proy:
        print(f"     ❌ Diferencia en cama: {mixto_cama_proy - mixto_calc} m³")
        discrepancias_encontradas = True
    else:
        print(f"     ✅ Mixto para cama coincide")

# Comparar arena
arena_calc = construccion_calc.get('ARENA M3')
arena_proy = construccion_proyecto.get('Arena Gruesa m3')

print(f"\n  ARENA:")
print(f"     Hoja de cálculos: {arena_calc} m³")
print(f"     Proyecto: {arena_proy} m³")
if arena_calc and arena_proy:
    if arena_calc != arena_proy:
        print(f"     ❌ Diferencia: {arena_proy - arena_calc} m³")
        discrepancias_encontradas = True
    else:
        print(f"     ✅ Arena coincide")

# Comparar malla
malla_calc = construccion_calc.get('MALLA')
malla_proy = construccion_proyecto.get('Malla sima 6 mm')

print(f"\n  MALLA:")
print(f"     Hoja de cálculos: {malla_calc} unidades")
print(f"     Proyecto: {malla_proy} unidades")
if malla_calc == malla_proy:
    print(f"     ✅ Malla coincide")
else:
    print(f"     ❌ Las cantidades difieren")
    discrepancias_encontradas = True

print("\n" + "=" * 100)
if discrepancias_encontradas:
    print("⚠️  SE ENCONTRARON DISCREPANCIAS ENTRE LAS DOS HOJAS")
    print("    Es necesario revisar y ajustar los valores para que coincidan.")
else:
    print("✅ TODAS LAS CANTIDADES PRINCIPALES COINCIDEN")
print("=" * 100)
